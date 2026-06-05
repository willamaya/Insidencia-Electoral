"""
6_reportes.py — Genera el Excel consolidado y el reporte final en Markdown
Genera:
  outputs/analisis_arrastre_boyaca_2026.xlsx  (5 hojas)
  outputs/resumen_modelo.md
"""
import logging
import pandas as pd
import numpy as np
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.settings import DATA_PROC, DATA_INTERIM, OUTPUTS, CONFIG

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# Colectividades con homologación CA↔SE real (excluye SIN_HOMOLOGACION y ALMA-OXIGENO individual)
COLECTIVIDADES_VALIDAS = [
    "ALIANZA VERDE", "PACTO HISTORICO", "CENTRO DEMOCRATICO",
    "PARTIDO LIBERAL", "CONSERVADOR-SALV NACIONAL",
    "CR-NUEVO LIBERALISMO", "PARTIDO DE LA U",
]


def hoja_resumen_boyaca(df_ca, df_se, idx_cand):
    """Resumen de votos totales y métricas clave por colectividad."""
    ca = (df_ca[df_ca["tipo_voto"]=="preferente"]
          .groupby("colectividad")
          .agg(votos_ca_total=("votos","sum"),
               candidatos_ca=("cedula","nunique"),
               mesas_ca=("amb_mesa","nunique")).reset_index())
    se = (df_se.groupby("colectividad")
          .agg(votos_se_total=("votos","sum")).reset_index())
    res = ca.merge(se, on="colectividad", how="left")
    res["ratio_arrastre"] = (res["votos_se_total"] / res["votos_ca_total"]).round(3)
    return res[res["colectividad"].isin(COLECTIVIDADES_VALIDAS)].sort_values("votos_ca_total", ascending=False)


def hoja_ranking_candidatos(idx_cand):
    """Top candidatos ordenados por índice ponderado, sólo col. válidas."""
    cols = ["candidato","colectividad","nombre_partido",
            "municipios","mesas_con_votos_ca",
            "votos_ca_total","votos_se_bruto_total",
            "indice_incidencia","indice_incidencia_pond","ratio_arrastre_global"]
    filt = idx_cand[idx_cand["colectividad"].isin(COLECTIVIDADES_VALIDAS)].copy()
    filt["indice_incidencia"]      = filt["indice_incidencia"].round(3)
    filt["indice_incidencia_pond"] = filt["indice_incidencia_pond"].round(3)
    filt["ratio_arrastre_global"]  = filt["ratio_arrastre_global"].round(3)
    return filt[cols].sort_values("indice_incidencia_pond", ascending=False)


def hoja_municipios(idx_muni):
    """Ratio de arrastre por municipio y colectividad."""
    filt = idx_muni[idx_muni["colectividad"].isin(COLECTIVIDADES_VALIDAS)].copy()
    filt["ratio_arrastre"] = filt["ratio_arrastre"].round(3)
    return filt.sort_values(["colectividad","ratio_arrastre"], ascending=[True,False])


def hoja_correlaciones(corr):
    """Tabla de correlaciones Pearson/Spearman."""
    return corr[corr["colectividad"].isin(COLECTIVIDADES_VALIDAS)].copy()


def hoja_regresion(reg):
    """Resultados OLS."""
    if reg is None or reg.empty:
        return pd.DataFrame({"nota":["statsmodels no disponible"]})
    filt = reg[reg["colectividad"].isin(COLECTIVIDADES_VALIDAS)].copy()
    filt["advertencia"] = "Datos agregados por mesa. NO inferir comportamiento individual."
    return filt


def escribir_excel(hojas: dict, path: Path):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        for nombre, df in hojas.items():
            df.to_excel(w, sheet_name=nombre[:31], index=False)
            ws = w.sheets[nombre[:31]]
            # Encabezado
            from openpyxl.styles import PatternFill, Font
            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = Font(color="FFFFFF", bold=True)
            # Autoajuste de columnas
            for col in ws.columns:
                max_w = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_w + 3, 50)
    log.info(f"Excel guardado: {path}")


def generar_md(idx_cand, corr, reg):
    """Resumen ejecutivo en Markdown."""
    top5 = idx_cand[idx_cand["colectividad"].isin(COLECTIVIDADES_VALIDAS)].head(5)
    lineas = [
        "# Análisis de Arrastre Cámara → Senado — Boyacá 2026\n",
        "## Metodología",
        "- **Unidad de análisis**: mesa electoral (3.282 mesas, Boyacá)",
        "- **Modelo determinístico**: ratio_arrastre = votos_SE_colectividad / votos_CA_partido por mesa",
        "- **Índice de incidencia ponderado**: Σ(votos_SE × dominancia_candidato) / Σ(votos_CA_candidato)",
        "- **Homologación**: cada codpar de Cámara mapeado a su equivalente de Senado",
        "- **⚠ Advertencia**: datos agregados por mesa, no inferencia individual (falacia ecológica)\n",
        "## Cobertura",
        "- Cámara: 3.282/3.282 mesas | 398 puestos | 123 municipios | 89.893 filas",
        "- Senado: 3.281/3.282 mesas | 398 puestos | 123 municipios | 130.732 filas\n",
        "## Top 5 candidatos por índice de incidencia ponderado",
        "| Candidato | Colectividad | Votos CA | Votos SE asoc. | Índice pond. | Mesas |",
        "|---|---|---|---|---|---|",
    ]
    for _, r in top5.iterrows():
        lineas.append(
            f"| {r['candidato']} | {r['colectividad']} | {r['votos_ca_total']:,} "
            f"| {int(r['votos_se_bruto_total']):,} | {r['indice_incidencia_pond']:.3f} "
            f"| {r['mesas_con_votos_ca']:,} |"
        )

    if corr is not None and not corr.empty:
        lineas += [
            "\n## Correlaciones Pearson (votos CA candidato vs votos SE colectividad, por mesa)",
            "| Colectividad | Pearson mesa | Pearson municipio |",
            "|---|---|---|",
        ]
        for _, r in corr[corr["colectividad"].isin(COLECTIVIDADES_VALIDAS)].iterrows():
            lineas.append(
                f"| {r['colectividad']} | {r['pearson_mesa']:.4f} | {r.get('pearson_municipio','N/A')} |"
            )

    if reg is not None and not reg.empty:
        lineas += [
            "\n## Regresión OLS (votos_SE ~ votos_CA + efectos_fijos_municipio)",
            "| Colectividad | Coef. votos_CA | p-valor | R² | Significativo |",
            "|---|---|---|---|---|",
        ]
        for _, r in reg[reg["colectividad"].isin(COLECTIVIDADES_VALIDAS)].iterrows():
            lineas.append(
                f"| {r['colectividad']} | {r['coef_votos_ca']:.4f} | {r['p_valor']:.4f} "
                f"| {r['r2']:.4f} | {'✓' if r['significativo_5pct'] else '✗'} |"
            )
        lineas.append(
            "\n> **Interpretación**: el coeficiente indica cuántos votos adicionales al Senado "
            "se asocian con 1 voto más a ese candidato a Cámara, controlando por municipio. "
            "**No implica causalidad individual.**"
        )

    return "\n".join(lineas)


def main():
    OUTPUTS.mkdir(parents=True, exist_ok=True)

    df_ca    = pd.read_parquet(DATA_INTERIM / "votos_camara_mesa.parquet")
    df_se    = pd.read_parquet(DATA_INTERIM / "votos_senado_mesa.parquet")
    idx_cand = pd.read_csv(OUTPUTS / "ranking_candidatos_incidencia.csv")
    idx_muni = pd.read_csv(OUTPUTS / "ranking_municipios_incidencia.csv")

    corr = pd.read_csv(OUTPUTS / "correlaciones.csv") if (OUTPUTS / "correlaciones.csv").exists() else None
    reg_path = OUTPUTS / "regresion_resultados.csv"
    reg  = pd.read_csv(reg_path) if reg_path.exists() else None

    # Excel
    hojas = {
        "Resumen Boyacá"    : hoja_resumen_boyaca(df_ca, df_se, idx_cand),
        "Ranking Candidatos": hoja_ranking_candidatos(idx_cand),
        "Por Municipio"     : hoja_municipios(idx_muni),
        "Correlaciones"     : hoja_correlaciones(corr) if corr is not None else pd.DataFrame(),
        "Regresión OLS"     : hoja_regresion(reg),
    }
    escribir_excel(hojas, OUTPUTS / "analisis_arrastre_boyaca_2026.xlsx")

    # Markdown
    md = generar_md(idx_cand, corr, reg)
    (OUTPUTS / "resumen_modelo.md").write_text(md, encoding="utf-8")
    log.info(f"Resumen Markdown: {OUTPUTS / 'resumen_modelo.md'}")

    # Imprimir resumen en consola
    log.info("\n" + md)


if __name__ == "__main__":
    main()
